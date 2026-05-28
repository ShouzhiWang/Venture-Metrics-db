from __future__ import annotations

import csv
import json
import re
import uuid
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook

from app.agents.demo_llm import DemoLLMClient, DemoLLMConfigError, DemoLLMProviderError, DemoLLMResponseError
from app.agents.query_planner import plan_query


TASK_TYPES = {
    "simple_answer",
    "find_data",
    "compare_definitions",
    "build_table",
    "create_excel",
    "aggregate_values",
    "source_audit",
    "source_comparison",
    "coverage_gap_analysis",
    "research_brief",
    "organization_mapping",
    "map_query",
}
NORMALIZED_COLUMNS = [
    "metric_name",
    "concept_group",
    "geography",
    "geography_match",
    "time_period",
    "value",
    "value_status",
    "unit",
    "dimension",
    "dimension_value",
    "source_report",
    "source_url",
    "availability",
    "evidence_quote",
    "confidence_score",
    "comparability_status",
    "notes",
]


@dataclass
class ResearchTaskPlan:
    query: str
    task_type: str
    domain: str | None = None
    geography: str | None = None
    time_range: str | None = None
    metric_type: str | None = None
    output_format: str | None = None
    unit_of_analysis: str | None = None
    availability: str | None = None
    comparison_target: str | None = None
    dimension: str | None = None
    dry_run: bool = False
    max_results: int = 30

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class ResearchTaskPlanner:
    def plan(self, query: str, context: dict[str, Any] | None = None, *, max_results: int = 30, dry_run: bool = False) -> ResearchTaskPlan:
        base = plan_query(query, context or {})
        detected = base.get("detected") or {}
        lowered = query.lower()
        task_type = self._task_type(lowered, detected)
        return ResearchTaskPlan(
            query=query.strip(),
            task_type=task_type,
            domain=detected.get("domain_topic"),
            geography=detected.get("geography"),
            time_range=detected.get("time_range"),
            metric_type=detected.get("metric_type"),
            output_format=self._output_format(task_type, detected),
            unit_of_analysis=detected.get("unit_of_analysis"),
            availability=detected.get("availability") or ("public_only" if "public" in lowered else None),
            comparison_target=detected.get("comparison_target"),
            dimension=self._dimension(lowered, detected),
            dry_run=dry_run,
            max_results=max_results,
        )

    def _task_type(self, lowered: str, detected: dict[str, Any]) -> str:
        if any(term in lowered for term in ("map ", "mapping", "ecosystem map")):
            return "organization_mapping" if any(term in lowered for term in ("organization", "ecosystem", "actor")) else "map_query"
        if re.search(r"\bgaps?\b", lowered) or re.search(r"\bcoverage\b", lowered):
            return "coverage_gap_analysis"
        if any(term in lowered for term in ("audit", "private sources", "source list", "what sources")):
            return "source_audit"
        if any(term in lowered for term in ("brief", "memo")):
            return "research_brief"
        if any(term in lowered for term in ("sum", "total", "aggregate", "add up")):
            return "aggregate_values"
        if "source comparison" in lowered or "compare sources" in lowered or "compare reports" in lowered:
            return "source_comparison"
        if any(term in lowered for term in ("compare", "definition", "definitions")):
            return "compare_definitions"
        if detected.get("output_format") == "excel" or any(term in lowered for term in ("excel", "xlsx", "workbook", "spreadsheet")):
            return "create_excel"
        if detected.get("output_format") == "table" or any(term in lowered for term in ("table", "csv", "dataset")):
            return "build_table"
        if any(term in lowered for term in ("find", "data", "metric", "indicator", "source")):
            return "find_data"
        return "simple_answer"

    def _output_format(self, task_type: str, detected: dict[str, Any]) -> str | None:
        if task_type == "create_excel":
            return "xlsx"
        if task_type == "build_table":
            return "csv"
        return detected.get("output_format")

    def _dimension(self, lowered: str, detected: dict[str, Any]) -> str | None:
        if "by stage" in lowered or "stage breakdown" in lowered:
            return "stage"
        if "by sector" in lowered or "sector breakdown" in lowered:
            return "sector"
        if "by country" in lowered:
            return "country"
        if detected.get("aggregation_intent") == "breakdown":
            return "breakdown"
        return None


def clarification_plan(query: str, context: dict[str, Any] | None = None) -> dict[str, Any] | None:
    plan = plan_query(query, context or {})
    if plan.get("action") != "ask_clarification":
        return None
    questions = plan.get("clarifying_questions") or []
    if not questions:
        return None
    return {
        "ok": True,
        "type": "clarification",
        "message": "Before I start the research task, I need one detail." if len(questions) == 1 else "Before I start the research task, I need a couple of details.",
        "query": query,
        "specificity": plan.get("specificity"),
        "intent": plan.get("intent"),
        "missing_dimensions": plan.get("missing_dimensions") or [],
        "clarifying_questions": questions[:2],
        "should_run_tool": False,
        "debug": {"clarification_plan": plan},
    }


class EvidencePacketBuilder:
    def build(self, query: str, task_plan: ResearchTaskPlan | dict[str, Any], retrieved: dict[str, Any]) -> dict[str, Any]:
        plan = task_plan if isinstance(task_plan, dict) else task_plan.to_dict()
        target_geo = plan.get("geography")
        variables = [self._variable(item, target_geo) for item in retrieved.get("closest_variables", []) or []]
        # Re-sort: exact_match first, then contextual, then mismatch, then unknown
        geo_order = {"exact_match": 0, "contextual_match": 1, "unknown": 2, "mismatch": 3}
        variables.sort(key=lambda v: (geo_order.get(v.get("geography_match", "unknown"), 2), -(v.get("confidence_score") or 0)))
        reports = [self._report(item) for item in retrieved.get("relevant_reports", []) or []]
        sources = [self._source(item) for item in retrieved.get("source_links", []) or []]
        organizations = [self._organization(item) for item in retrieved.get("relevant_organizations", []) or []]
        # Geography mismatch warning
        geo_warnings = []
        if target_geo:
            mismatched = [v for v in variables if v.get("geography_match") == "mismatch"]
            if mismatched and not any(v.get("geography_match") == "exact_match" for v in variables):
                mismatched_geos = sorted({v.get("geography") for v in mismatched if v.get("geography")})
                geo_warnings.append(f"No exact {target_geo} data found. Showing data from: {', '.join(mismatched_geos[:3])}.")
        return {
            "query": query,
            "interpreted_intent": plan,
            "variables": variables,
            "reports": reports,
            "sources": sources,
            "organizations": organizations,
            "evidence_quotes": [item.get("evidence_quote") for item in variables if item.get("evidence_quote")],
            "source_urls": sorted({item.get("source_url") for item in variables + reports + sources if item.get("source_url")}),
            "availability_labels": sorted({item.get("availability") for item in variables + sources if item.get("availability")}),
            "geography_coverage": sorted({item.get("geography") for item in variables + reports if item.get("geography")}),
            "time_coverage": sorted({item.get("time_period") for item in variables if item.get("time_period")}),
            "confidence_scores": [item.get("confidence_score") for item in variables if item.get("confidence_score") is not None],
            "limitations": list(retrieved.get("limitations") or []) + geo_warnings,
        }

    def _variable(self, item: dict[str, Any], target_geography: str | None = None) -> dict[str, Any]:
        structured = structured_value_fields(item)
        value = structured["value"]
        unit = structured["unit"]
        value_status = structured["value_status"]
        item_geo = first_present(item, "geographic_coverage", "geography")
        return {
            "id": item.get("object_id") or item.get("id") or item.get("variable_id"),
            "metric_name": item.get("title") or item.get("raw_variable_name"),
            "concept_group": concept_group(item),
            "definition": item.get("definition"),
            "measurement_method": item.get("measurement_method"),
            "geography": item_geo,
            "geography_match": classify_geography_match(item_geo, target_geography),
            "time_period": first_present(item, "temporal_coverage", "time_period", "time_coverage"),
            "value": value,
            "value_status": value_status,
            "unit": item.get("unit") or unit,
            "dimension": infer_dimension(item),
            "dimension_value": infer_dimension_value(item),
            "source_report": item.get("report_title") or item.get("source_report") or item.get("data_source"),
            "source_url": item.get("source_url"),
            "availability": item.get("availability") or "unclear",
            "evidence_quote": item.get("evidence_quote") or item.get("why_it_matched"),
            "confidence_score": item.get("confidence_score") or item.get("score"),
            "directness": "direct" if item.get("score", 0) and item.get("score", 0) >= 0.5 else "contextual",
            "notes": item.get("why_it_matched"),
        }

    def _report(self, item: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": item.get("object_id") or item.get("id") or item.get("report_id"),
            "title": item.get("title"),
            "publisher": item.get("publisher"),
            "geography": item.get("geographic_coverage") or item.get("geography"),
            "time_period": item.get("temporal_coverage") or item.get("report_year"),
            "source_url": item.get("source_url"),
            "availability": item.get("availability") or "unclear",
            "confidence_score": item.get("score"),
            "evidence_quote": item.get("evidence_quote") or item.get("why_it_matched"),
        }

    def _source(self, item: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": item.get("object_id") or item.get("source_id"),
            "title": item.get("title"),
            "source_url": item.get("source_url"),
            "availability": item.get("availability") or "unclear",
            "local_path": item.get("local_path"),
            "confidence_score": item.get("score"),
        }

    def _organization(self, item: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": item.get("object_id") or item.get("id"),
            "name": item.get("name") or item.get("title"),
            "organization_type": item.get("organization_type"),
            "geography": item.get("geography"),
            "website_url": item.get("website_url") or item.get("source_url"),
            "confidence_score": item.get("score"),
        }


class ComparabilityValidator:
    def validate(self, rows: list[dict[str, Any]], *, aggregation_requested: bool = False) -> dict[str, Any]:
        if not rows:
            return {
                "status": "insufficient_metadata",
                "issues": ["No rows available for comparison."],
                "issue_details": [{"code": "no_rows", "field": None, "message": "No rows available for comparison."}],
                "can_aggregate": False,
                "explanation": "Aggregation is blocked because there are no retrieved rows to compare.",
                "safe_aggregation_requirements": safe_aggregation_requirements(),
                "comparison_table": [],
            }
        issue_details: list[dict[str, Any]] = []
        for field_name, label in [
            ("geography", "geography"),
            ("time_period", "time period"),
            ("unit", "unit"),
            ("metric_name", "metric definition/name"),
            ("dimension", "dimension"),
        ]:
            values = {normalize_blank(row.get(field_name)) for row in rows}
            values.discard("")
            if len(values) > 1:
                issue_details.append(
                    {
                        "code": f"{field_name}_mismatch",
                        "field": field_name,
                        "message": f"Mixed {label}: {', '.join(sorted(values))}.",
                        "values": sorted(values),
                    }
                )
            elif not values:
                issue_details.append(
                    {
                        "code": f"missing_{field_name}",
                        "field": field_name,
                        "message": f"Missing {label} metadata.",
                        "values": [],
                    }
                )
        availability = {normalize_blank(row.get("availability")) for row in rows}
        if "private" in availability or "not_obtainable" in availability:
            issue_details.append(
                {
                    "code": "private_source_limitation",
                    "field": "availability",
                    "message": "Private or not-obtainable sources limit reproducible aggregation.",
                    "values": sorted(value for value in availability if value),
                }
            )
        source_urls = [row.get("source_url") for row in rows if row.get("source_url")]
        if len(source_urls) != len(set(source_urls)):
            issue_details.append(
                {
                    "code": "source_overlap_double_counting",
                    "field": "source_url",
                    "message": "Source overlap creates double-counting risk.",
                    "values": sorted({str(url) for url in source_urls if source_urls.count(url) > 1}),
                }
            )
        issues = [item["message"] for item in issue_details]
        status = self._status(issues)
        can_aggregate = aggregation_requested and status == "comparable"
        return {
            "status": status,
            "issues": issues,
            "issue_details": issue_details,
            "can_aggregate": can_aggregate,
            "explanation": comparability_explanation(status, issue_details, aggregation_requested),
            "safe_aggregation_requirements": safe_aggregation_requirements(),
            "comparison_table": comparability_table(rows, issue_details),
        }

    def _status(self, issues: list[str]) -> str:
        if not issues:
            return "comparable"
        metadata_issues = [item for item in issues if item.startswith("Missing")]
        if len(metadata_issues) == len(issues):
            return "insufficient_metadata"
        if any("Mixed unit" in item or "double-counting" in item or "Private" in item for item in issues):
            return "not_comparable"
        return "partially_comparable"


class AnswerSynthesizer:
    def __init__(self, llm_client: Any | None = None, *, use_llm: bool = False) -> None:
        self.llm_client = llm_client
        self.use_llm = use_llm

    def synthesize(self, evidence_packet: dict[str, Any], comparability: dict[str, Any] | None = None) -> str:
        if self.use_llm or self.llm_client is not None:
            try:
                client = self.llm_client or DemoLLMClient()
                if hasattr(client, "synthesize_research_task"):
                    answer = client.synthesize_research_task(evidence_packet=evidence_packet, comparability=comparability or {})
                else:
                    answer = client._text_completion(self.prompt_text(evidence_packet, comparability or {}))
                if isinstance(answer, str) and answer.strip():
                    return answer.strip()
            except (DemoLLMConfigError, DemoLLMProviderError, DemoLLMResponseError, AttributeError):
                pass
        return self._fallback_synthesize(evidence_packet, comparability)

    def _fallback_synthesize(self, evidence_packet: dict[str, Any], comparability: dict[str, Any] | None = None) -> str:
        variables = evidence_packet.get("variables") or []
        reports = evidence_packet.get("reports") or []
        sources = evidence_packet.get("sources") or []
        query = evidence_packet.get("query") or "this request"
        plan = evidence_packet.get("interpreted_intent") or {}
        target_geo = plan.get("geography")

        if not variables and not reports and not sources:
            return self._empty_result_response(query, target_geo, evidence_packet)
        direct = [item for item in variables if item.get("directness") == "direct"]
        contextual = [item for item in variables if item.get("directness") != "direct"]
        lines = [self._direct_answer(evidence_packet, direct, variables)]
        grouped = group_by_concept(variables)
        if grouped:
            lines.append("Concept groups: " + "; ".join(f"{name} ({len(items)})" for name, items in grouped.items()) + ".")
        if direct:
            lines.append("Direct matches: " + "; ".join(self._variable_summary(item) for item in direct[:5]) + ".")
        if contextual:
            lines.append("Contextual matches: " + "; ".join(self._variable_summary(item) for item in contextual[:5]) + ".")
        cited_reports = [item.get("title") for item in reports if item.get("title")]
        cited_sources = [item.get("title") or item.get("source_url") for item in sources if item.get("title") or item.get("source_url")]
        if cited_reports or cited_sources:
            lines.append("Cited evidence: " + "; ".join((cited_reports + cited_sources)[:6]) + ".")
        labels = evidence_packet.get("availability_labels") or []
        if labels:
            lines.append("Availability: " + ", ".join(labels) + ".")
        limitations = list(evidence_packet.get("limitations") or [])
        if comparability and comparability.get("issues"):
            limitations.extend(comparability["issues"])
        if limitations:
            lines.append("Limitations: " + " ".join(limitations))
        if comparability and comparability.get("status") not in {None, "comparable"}:
            lines.append((comparability.get("explanation") or "").strip())
            requirements = comparability.get("safe_aggregation_requirements") or []
            if requirements:
                lines.append("To aggregate safely, the rows need " + "; ".join(requirements[:6]) + ".")
        lines.append("Next actions: export the evidence table, inspect source reports, or refine by geography, time period, metric type, or data availability.")
        return "\n\n".join(lines)

    def prompt_payload(self, evidence_packet: dict[str, Any]) -> dict[str, Any]:
        return {
            "instruction": (
                "Lead with a direct answer. Distinguish direct and contextual matches. "
                "Group variables into concepts, extract values/units/years only when present, cite source/report names, "
                "label availability, explain limitations, then suggest next actions."
            ),
            "evidence_packet": evidence_packet,
        }

    def prompt_text(self, evidence_packet: dict[str, Any], comparability: dict[str, Any]) -> str:
        payload = self.prompt_payload(evidence_packet)
        return (
            payload["instruction"]
            + "\n\nUse only this evidence packet and comparability result. Do not invent values.\n\n"
            + json.dumps({"evidence_packet": evidence_packet, "comparability": comparability}, ensure_ascii=True, default=str)
        )

    def _direct_answer(self, evidence_packet: dict[str, Any], direct: list[dict[str, Any]], variables: list[dict[str, Any]]) -> str:
        query = evidence_packet.get("query") or "this request"
        count = len(variables)
        direct_count = len(direct)
        if direct_count:
            return f"Direct answer: I found {direct_count} direct variable match{'es' if direct_count != 1 else ''} and {count - direct_count} contextual match{'es' if count - direct_count != 1 else ''} for {query}."
        return f"Direct answer: I found {count} contextual variable match{'es' if count != 1 else ''} for {query}, but no high-confidence direct variable match."

    def _variable_summary(self, item: dict[str, Any]) -> str:
        bits = [str(item.get("metric_name") or "Unnamed metric")]
        if item.get("value") is not None:
            unit = f" {item.get('unit')}" if item.get("unit") else ""
            bits.append(f"value {item['value']}{unit}")
        if item.get("time_period"):
            bits.append(str(item["time_period"]))
        if item.get("geography"):
            bits.append(str(item["geography"]))
        if item.get("availability"):
            bits.append(str(item["availability"]))
        return " / ".join(bits)

    def _empty_result_response(self, query: str, target_geo: str | None, evidence_packet: dict[str, Any]) -> str:
        """Generate a user-friendly response when no results are found."""
        lines = [f'I could not find matching data for "{query}" in the current database.']
        # Say what was searched
        search_parts = []
        if target_geo:
            search_parts.append(f"geography: {target_geo}")
        plan = evidence_packet.get("interpreted_intent") or {}
        if plan.get("domain"):
            search_parts.append(f"domain: {plan['domain']}")
        if plan.get("metric_type"):
            search_parts.append(f"metric: {plan['metric_type']}")
        if search_parts:
            lines.append(f"Searched for: {', '.join(search_parts)}.")
        # Say what's missing
        lines.append("The database does not yet contain matching variables, reports, or sources for this specific query.")
        # Suggest next actions
        suggestions = []
        if target_geo:
            suggestions.append(f"Try broadening the geography (remove '{target_geo}' filter)")
        suggestions.append("Search for related reports only")
        suggestions.append("Search for ecosystem organizations in this area")
        suggestions.append("Include private/proprietary sources in results")
        suggestions.append("Check if data exists under a different name (e.g., 'venture capital' instead of 'startup funding')")
        lines.append("You can try:")
        for s in suggestions[:5]:
            lines.append(f"  - {s}")
        return "\n".join(lines)


class TableExcelExportService:
    def build_rows(self, evidence_packet: dict[str, Any], comparability: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        status = (comparability or {}).get("status") or "insufficient_metadata"
        rows = []
        for item in evidence_packet.get("variables") or []:
            rows.append(
                {
                    "metric_name": item.get("metric_name"),
                    "concept_group": item.get("concept_group"),
                    "geography": item.get("geography"),
                    "time_period": item.get("time_period"),
                    "value": item.get("value"),
                    "value_status": item.get("value_status"),
                    "unit": item.get("unit"),
                    "dimension": item.get("dimension"),
                    "dimension_value": item.get("dimension_value"),
                    "source_report": item.get("source_report"),
                    "source_url": item.get("source_url"),
                    "availability": item.get("availability"),
                    "evidence_quote": item.get("evidence_quote"),
                    "confidence_score": item.get("confidence_score"),
                    "comparability_status": status,
                    "notes": item.get("notes"),
                }
            )
        return rows

    def export(
        self,
        evidence_packet: dict[str, Any],
        *,
        output_dir: str | Path,
        output_format: str,
        comparability: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        stem = f"research_task_{uuid.uuid4().hex[:8]}"
        rows = self.build_rows(evidence_packet, comparability)
        all_paths = []
        if output_format == "xlsx":
            path = output_path / f"{stem}.xlsx"
            self._write_xlsx(path, evidence_packet, rows, comparability or {})
            all_paths.append(str(path))
        elif output_format == "csv":
            # Write main data file + sidecar files for metadata
            data_path = output_path / f"{stem}_data.csv"
            self._write_csv(data_path, rows)
            all_paths.append(str(data_path))
            # Sidecar: methodology notes
            notes_path = output_path / f"{stem}_methodology_notes.csv"
            notes = methodology_notes(evidence_packet, comparability or {})
            self._write_sidecar_csv(notes_path, ["note_type", "note"], notes)
            all_paths.append(str(notes_path))
            # Sidecar: data gaps
            gaps_path = output_path / f"{stem}_data_gaps.csv"
            gaps = data_gaps(evidence_packet, comparability or {})
            self._write_sidecar_csv(gaps_path, ["gap_type", "description"], gaps)
            all_paths.append(str(gaps_path))
            # Sidecar: source reports
            reports_path = output_path / f"{stem}_source_reports.csv"
            report_cols = ["id", "title", "publisher", "geography", "time_period", "source_url", "availability", "confidence_score", "evidence_quote"]
            self._write_sidecar_csv(reports_path, report_cols, evidence_packet.get("reports") or [])
            all_paths.append(str(reports_path))
            path = data_path  # Primary path for backward compat
        elif output_format == "json":
            path = output_path / f"{stem}.json"
            path.write_text(json.dumps({"evidence_packet": evidence_packet, "normalized_data": rows, "comparability": comparability}, indent=2, default=str), encoding="utf-8")
            all_paths.append(str(path))
        else:
            raise ValueError("output_format must be xlsx, csv, or json.")
        return {"path": str(path), "format": output_format, "row_count": len(rows), "all_paths": all_paths}

    def _write_csv(self, path: Path, rows: list[dict[str, Any]]) -> None:
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=NORMALIZED_COLUMNS)
            writer.writeheader()
            writer.writerows([{column: row.get(column) for column in NORMALIZED_COLUMNS} for row in rows])

    def _write_sidecar_csv(self, path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows([{col: row.get(col) for col in columns} for row in rows])

    def build_source_comparison(self, evidence_packet: dict[str, Any]) -> list[dict[str, Any]]:
        """Build a side-by-side source comparison table."""
        reports = evidence_packet.get("reports") or []
        variables = evidence_packet.get("variables") or []
        # Group variables by source report
        vars_by_report: dict[str, list[dict[str, Any]]] = {}
        for v in variables:
            report = v.get("source_report") or "Unknown"
            vars_by_report.setdefault(report, []).append(v)
        # If no reports, build from variable source_report groups
        if not reports and vars_by_report:
            comparison = []
            for report_name, report_vars in vars_by_report.items():
                geos = sorted({v.get("geography") for v in report_vars if v.get("geography")})
                times = sorted({v.get("time_period") for v in report_vars if v.get("time_period")})
                avail = sorted({v.get("availability") for v in report_vars if v.get("availability")})
                concept_groups = sorted({v.get("concept_group") for v in report_vars if v.get("concept_group")})
                values_with_data = [v for v in report_vars if v.get("value") is not None]
                comparison.append({
                    "source_title": report_name[:80],
                    "publisher": "",
                    "geography": "; ".join(geos[:3]) if geos else "",
                    "time_coverage": "; ".join(times[:3]) if times else "",
                    "variables_found": len(report_vars),
                    "values_extracted": len(values_with_data),
                    "concept_groups": "; ".join(concept_groups[:3]) if concept_groups else "",
                    "availability": "; ".join(avail[:3]) if avail else "",
                    "source_url": "",
                    "confidence": "",
                })
            return comparison
        if not reports:
            return []
        comparison = []
        for report in reports:
            title = report.get("title") or "Unknown"
            # Find matching variables
            report_vars = vars_by_report.get(title, [])
            # Also try partial match
            if not report_vars:
                for key, vals in vars_by_report.items():
                    if key and title and (key[:30] in title or title[:30] in key):
                        report_vars = vals
                        break
            geos = sorted({v.get("geography") for v in report_vars if v.get("geography")})
            times = sorted({v.get("time_period") for v in report_vars if v.get("time_period")})
            avail = sorted({v.get("availability") for v in report_vars if v.get("availability")})
            concept_groups = sorted({v.get("concept_group") for v in report_vars if v.get("concept_group")})
            values_with_data = [v for v in report_vars if v.get("value") is not None]
            comparison.append({
                "source_title": title[:80],
                "publisher": report.get("publisher") or "",
                "geography": "; ".join(geos[:3]) if geos else (report.get("geography") or ""),
                "time_coverage": "; ".join(times[:3]) if times else (report.get("time_period") or ""),
                "variables_found": len(report_vars),
                "values_extracted": len(values_with_data),
                "concept_groups": "; ".join(concept_groups[:3]) if concept_groups else "",
                "availability": "; ".join(avail[:3]) if avail else (report.get("availability") or ""),
                "source_url": report.get("source_url") or "",
                "confidence": report.get("confidence_score") or "",
            })
        return comparison

    def _write_xlsx(
        self,
        path: Path,
        evidence_packet: dict[str, Any],
        rows: list[dict[str, Any]],
        comparability: dict[str, Any],
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "normalized_data"
        append_table(ws, NORMALIZED_COLUMNS, rows)
        append_table(wb.create_sheet("source_variables"), sorted_variable_columns(evidence_packet), evidence_packet.get("variables") or [])
        append_table(wb.create_sheet("source_reports"), ["id", "title", "publisher", "geography", "time_period", "source_url", "availability", "confidence_score", "evidence_quote"], evidence_packet.get("reports") or [])
        notes = methodology_notes(evidence_packet, comparability)
        append_table(wb.create_sheet("methodology_notes"), ["note_type", "note"], notes)
        append_table(wb.create_sheet("data_gaps"), ["gap_type", "description"], data_gaps(evidence_packet, comparability))
        # Source comparison sheet
        source_comp = self.build_source_comparison(evidence_packet)
        if source_comp:
            comp_cols = ["source_title", "publisher", "geography", "time_coverage", "variables_found", "values_extracted", "concept_groups", "availability", "source_url", "confidence"]
            append_table(wb.create_sheet("source_comparison"), comp_cols, source_comp)
        wb.save(path)


def execute_research_task(
    query: str,
    *,
    tool_caller: Callable[[str, dict[str, Any]], dict[str, Any]],
    context: dict[str, Any] | None = None,
    output_dir: str | Path = "exports/research_tasks",
    output_format: str | None = None,
    dry_run: bool = False,
    max_results: int = 30,
    llm_client: Any | None = None,
    use_llm: bool = True,
    run_with_defaults: bool = False,
) -> dict[str, Any]:
    clarification = clarification_plan(query, context)
    if clarification and not run_with_defaults:
        # Add run_with_defaults option for export/dataset tasks
        intent = clarification.get("intent", "")
        if intent in ("create_excel", "build_table"):
            clarification["run_with_defaults_option"] = {
                "message": "Or run with defaults: all available geographies and time periods, public and private sources labeled.",
                "defaults": {"geography": "all available", "time_range": "all available", "availability": "all"},
            }
        return clarification
    planner = ResearchTaskPlanner()
    task_plan = planner.plan(query, context, max_results=max_results, dry_run=dry_run)
    # If run_with_defaults, override missing geography/time with None (search all)
    if run_with_defaults:
        task_plan.geography = task_plan.geography or None
        task_plan.availability = None  # Include all
    retrieve_args = {
        "query": query,
        "limit": min(max_results, 25),
        "public_only": task_plan.availability == "public_only",
        "geography": task_plan.geography,
        "time_range": task_plan.time_range,
    }
    retrieved = {"closest_variables": [], "relevant_reports": [], "source_links": [], "relevant_organizations": [], "limitations": []}
    if not dry_run:
        if task_plan.task_type == "organization_mapping":
            tool_result = tool_caller("semantic_search", {"query": query, "object_types": ["organization"], "limit": min(max_results, 25)})
            retrieved = normalize_semantic_results(tool_result)
        elif task_plan.task_type == "compare_definitions":
            tool_result = tool_caller("compare_concepts_auto", {"query": query, "geography": task_plan.geography, "public_only": task_plan.availability == "public_only"})
            retrieved = normalize_compare_results(tool_result)
        else:
            tool_result = tool_caller("find_data", {key: value for key, value in retrieve_args.items() if value is not None})
            retrieved = normalize_find_data_results(tool_result)
    packet = EvidencePacketBuilder().build(query, task_plan, retrieved)
    if run_with_defaults:
        packet.setdefault("limitations", []).append("Running with defaults: all available geographies and time periods. Results may not be directly comparable.")
    export_service = TableExcelExportService()
    rows = export_service.build_rows(packet)
    comparability = ComparabilityValidator().validate(rows, aggregation_requested=task_plan.task_type == "aggregate_values")
    packet["comparability"] = comparability
    # Build source comparison if task type is source_comparison
    source_comparison = None
    if task_plan.task_type == "source_comparison":
        source_comparison = export_service.build_source_comparison(packet)
        packet["source_comparison"] = source_comparison
    answer = AnswerSynthesizer(llm_client=llm_client, use_llm=use_llm).synthesize(packet, comparability)
    # If aggregation is blocked and we have a comparison table, append it to the answer
    if comparability.get("status") not in (None, "comparable") and comparability.get("comparison_table"):
        table = comparability["comparison_table"]
        if table and "comparison table" not in answer.lower():
            answer += "\n\n### Comparison Table\n\n"
            answer += "| Metric | Value | Unit | Geography | Time | Source | Why not comparable | Safe use |\n"
            answer += "|--------|-------|------|-----------|------|--------|-------------------|----------|\n"
            for row in table[:15]:
                vals = [
                    str(row.get("metric_name") or "")[:30],
                    str(row.get("value") or ""),
                    str(row.get("unit") or "")[:15],
                    str(row.get("geography") or "")[:15],
                    str(row.get("time_period") or "")[:15],
                    str(row.get("source_report") or "")[:25],
                    str(row.get("why_not_comparable") or "")[:30],
                    str(row.get("safe_use") or "")[:20],
                ]
                answer += "| " + " | ".join(vals) + " |\n"
    export = None
    requested_format = output_format or task_plan.output_format
    # Trigger export for source_comparison tasks too
    export_task_types = {"build_table", "create_excel", "aggregate_values", "source_comparison"}
    if not dry_run and requested_format in {"xlsx", "csv", "json"} and task_plan.task_type in export_task_types:
        export = export_service.export(packet, output_dir=output_dir, output_format=requested_format, comparability=comparability)
    return {
        "ok": True,
        "task_plan": task_plan.to_dict(),
        "evidence_packet": packet,
        "comparability": comparability,
        "source_comparison": source_comparison,
        "answer": answer,
        "normalized_data": rows,
        "export": export,
        "dry_run": dry_run,
    }


def normalize_find_data_results(tool_result: dict[str, Any]) -> dict[str, Any]:
    if not tool_result.get("ok"):
        return {"closest_variables": [], "relevant_reports": [], "source_links": [], "relevant_organizations": [], "limitations": [tool_result.get("error", {}).get("message", "Tool failed.")]}
    data = tool_result.get("data") or {}
    return {
        "closest_variables": data.get("closest_variables") or [],
        "relevant_reports": data.get("relevant_reports") or [],
        "source_links": data.get("source_links") or [],
        "relevant_organizations": data.get("relevant_organizations") or [],
        "limitations": [item for item in [data.get("warning")] if item],
    }


def normalize_compare_results(tool_result: dict[str, Any]) -> dict[str, Any]:
    if not tool_result.get("ok"):
        return normalize_find_data_results(tool_result)
    data = tool_result.get("data") or {}
    return {
        "closest_variables": data.get("closest_variables") or [],
        "relevant_reports": data.get("selected_reports") or [],
        "source_links": [],
        "relevant_organizations": [],
        "limitations": data.get("limitations") or [],
    }


def normalize_semantic_results(tool_result: dict[str, Any]) -> dict[str, Any]:
    if not tool_result.get("ok"):
        return normalize_find_data_results(tool_result)
    rows = (tool_result.get("data") or {}).get("results") or []
    organizations = []
    sources = []
    for row in rows:
        if row.get("object_type") == "organization":
            metadata = row.get("metadata") or {}
            organizations.append(
                {
                    "object_id": row.get("object_id"),
                    "name": row.get("title"),
                    "organization_type": metadata.get("organization_type"),
                    "geography": row.get("geography"),
                    "website_url": metadata.get("website_url") or row.get("source_url"),
                    "score": row.get("score"),
                }
            )
        elif row.get("source_url"):
            sources.append({"title": row.get("title"), "source_url": row.get("source_url"), "availability": row.get("availability"), "score": row.get("score")})
    return {"closest_variables": [], "relevant_reports": [], "source_links": sources, "relevant_organizations": organizations, "limitations": []}


def extract_numeric_value(text: str) -> tuple[float | None, str | None]:
    if not text:
        return None, None
    match = re.search(r"(?P<prefix>[$€£])?\b(?P<number>\d+(?:,\d{3})*(?:\.\d+)?)\s*(?P<suffix>%|percent|million|billion|mn|bn)?", text, re.I)
    if not match:
        return None, None
    number = float(match.group("number").replace(",", ""))
    suffix = match.group("suffix")
    prefix = match.group("prefix")
    unit = prefix or suffix
    return number, unit


def structured_value_fields(item: dict[str, Any]) -> dict[str, Any]:
    metadata = item.get("metadata") if isinstance(item.get("metadata"), dict) else {}
    value = first_present(item, "value", "metric_value", "numeric_value", "amount")
    unit = first_present(item, "unit", "value_unit", "currency", "measurement_unit")
    if value is None and metadata:
        value = first_present(metadata, "value", "metric_value", "numeric_value", "amount")
    if unit is None and metadata:
        unit = first_present(metadata, "unit", "value_unit", "currency", "measurement_unit")
    parsed_value = parse_number(value)
    if parsed_value is not None:
        return {"value": parsed_value, "unit": unit, "value_status": "structured"}
    fallback_value, fallback_unit = extract_numeric_value(item.get("evidence_quote") or item.get("why_it_matched") or "")
    if fallback_value is not None:
        return {"value": fallback_value, "unit": unit or fallback_unit, "value_status": "extracted_from_evidence"}
    return {"value": None, "unit": unit, "value_status": "not_extracted"}


def first_present(mapping: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = mapping.get(key)
        if value not in (None, ""):
            return value
    return None


def parse_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.strip():
        match = re.search(r"-?\d+(?:,\d{3})*(?:\.\d+)?", value)
        if match:
            return float(match.group(0).replace(",", ""))
    return None


def concept_group(item: dict[str, Any]) -> str | None:
    text = " ".join(str(item.get(key) or "") for key in ("title", "definition", "measurement_method")).lower()
    if any(term in text for term in ("stage", "seed", "series a")):
        return "stage breakdown"
    if "sector" in text:
        return "sector breakdown"
    if any(term in text for term in ("deal", "round")):
        return "deal activity"
    if any(term in text for term in ("funding", "investment", "capital")):
        return "funding amount"
    if any(term in text for term in ("exit", "ipo", "acquisition")):
        return "exits"
    return item.get("concept_group") or "other"


def infer_dimension(item: dict[str, Any]) -> str | None:
    text = " ".join(str(item.get(key) or "") for key in ("title", "definition", "measurement_method")).lower()
    if "stage" in text or "seed" in text or "series a" in text:
        return "stage"
    if "sector" in text:
        return "sector"
    if "country" in text:
        return "country"
    return None


def infer_dimension_value(item: dict[str, Any]) -> str | None:
    text = " ".join(str(item.get(key) or "") for key in ("title", "definition")).lower()
    for value in ("pre-seed", "seed", "series a", "series b", "growth", "late stage"):
        if value in text:
            return value.title()
    return None


def group_by_concept(variables: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in variables:
        grouped.setdefault(item.get("concept_group") or "other", []).append(item)
    return grouped


def normalize_blank(value: Any) -> str:
    return str(value or "").strip().lower()


def classify_geography_match(item_geography: str | None, target_geography: str | None) -> str:
    """Classify how well an item's geography matches the target."""
    if not target_geography:
        return "unknown"
    if not item_geography:
        return "unknown"
    item_lower = normalize_blank(item_geography)
    target_lower = normalize_blank(target_geography)
    # Exact match
    if item_lower == target_lower:
        return "exact_match"
    # Target is contained in item (e.g., "Singapore" in "Singapore and ASEAN 6")
    if target_lower in item_lower:
        return "contextual_match"
    # Item is contained in target (e.g., "ASEAN" in "Southeast Asia")
    if item_lower in target_lower:
        return "contextual_match"
    # Known related geographies
    geo_groups = [
        {"singapore", "asean", "southeast asia", "sea"},
        {"hong kong", "china", "greater bay area", "gba"},
        {"asia", "asian", "southeast asia", "east asia", "south asia"},
    ]
    for group in geo_groups:
        if target_lower in group and item_lower in group:
            return "contextual_match"
    return "mismatch"


def comparability_table(rows: list[dict[str, Any]], issue_details: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    issue_fields = set()
    if issue_details:
        for item in issue_details:
            field = item.get("field")
            if field:
                issue_fields.add(field)
    table = []
    for row in rows:
        # Determine why this row is not comparable
        why_parts = []
        if "geography" in issue_fields:
            why_parts.append(f"geography: {row.get('geography') or 'missing'}")
        if "time_period" in issue_fields:
            why_parts.append(f"time: {row.get('time_period') or 'missing'}")
        if "unit" in issue_fields:
            why_parts.append(f"unit: {row.get('unit') or 'missing'}")
        if "metric_name" in issue_fields:
            why_parts.append("different metric")
        # Determine safe use
        safe_parts = []
        if row.get("geography") and "geography" not in issue_fields:
            safe_parts.append("same geography")
        if row.get("time_period") and "time_period" not in issue_fields:
            safe_parts.append("same time period")
        if row.get("availability") in ("obtainable", "public"):
            safe_parts.append("public data")
        table.append({
            "metric_name": row.get("metric_name"),
            "value": row.get("value"),
            "unit": row.get("unit"),
            "geography": row.get("geography"),
            "time_period": row.get("time_period"),
            "source_report": row.get("source_report"),
            "source_url": row.get("source_url"),
            "availability": row.get("availability"),
            "why_not_comparable": "; ".join(why_parts) if why_parts else "",
            "safe_use": "; ".join(safe_parts) if safe_parts else "use with caution",
        })
    return table


def comparability_explanation(status: str, issue_details: list[dict[str, Any]], aggregation_requested: bool) -> str:
    if status == "comparable":
        return "These rows measure the same thing, from the same kind of source, over the same time period. They can be combined."
    if not issue_details:
        if aggregation_requested:
            return "I cannot add these values because I do not have enough information about what each source measures."
        return "I do not have enough information to compare these rows."
    plain_reasons = []
    for item in issue_details:
        code = item.get("code")
        if code == "unit_mismatch":
            plain_reasons.append("they use different units (e.g., count vs. USD vs. percentage)")
        elif code == "geography_mismatch":
            plain_reasons.append("they cover different countries or regions")
        elif code == "time_period_mismatch":
            plain_reasons.append("they cover different years or time periods")
        elif code == "metric_name_mismatch":
            plain_reasons.append("they measure different things (e.g., deal value vs. IPO proceeds vs. funding amount)")
        elif code == "dimension_mismatch":
            plain_reasons.append("they break down data differently (e.g., by stage vs. by sector)")
        elif code == "source_overlap_double_counting":
            plain_reasons.append("some sources overlap, which would double-count the same data")
        elif code == "private_source_limitation":
            plain_reasons.append("some sources are private, so the combined result would not be reproducible")
        elif str(code).startswith("missing_"):
            plain_reasons.append(f"some rows are missing {item.get('field', 'key')} information")
        else:
            plain_reasons.append(item.get("message", "there is a data quality issue").rstrip(".").lower())
    if aggregation_requested:
        intro = "I should not add these into one total"
    else:
        intro = "These rows are not directly comparable"
    reason_text = "; ".join(plain_reasons)
    return f"{intro} because {reason_text}. I created a comparison table instead so you can see each row side by side."


def safe_aggregation_requirements() -> list[str]:
    return [
        "same geography",
        "same time period",
        "same unit/currency",
        "same metric definition and funding type",
        "same dimension and dimension value",
        "non-overlapping source coverage",
        "clear public/obtainable availability or explicit private-source caveat",
    ]


def append_table(ws, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([_sanitize_cell(row.get(column)) for column in columns])


def _sanitize_cell(value: Any) -> Any:
    """Remove control characters that openpyxl can't handle in worksheets."""
    if isinstance(value, str):
        # Remove control chars except tab, newline, carriage return
        return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    return value


def sorted_variable_columns(evidence_packet: dict[str, Any]) -> list[str]:
    preferred = [
        "id",
        "metric_name",
        "concept_group",
        "definition",
        "measurement_method",
        "geography",
        "time_period",
        "value",
        "value_status",
        "unit",
        "dimension",
        "dimension_value",
        "source_report",
        "source_url",
        "availability",
        "evidence_quote",
        "confidence_score",
        "directness",
        "notes",
    ]
    keys = set()
    for item in evidence_packet.get("variables") or []:
        keys.update(item)
    return preferred + sorted(keys - set(preferred))


def methodology_notes(evidence_packet: dict[str, Any], comparability: dict[str, Any]) -> list[dict[str, str]]:
    plan = evidence_packet.get("interpreted_intent") or {}
    notes = [
        {"note_type": "query", "note": str(evidence_packet.get("query") or "")},
        {"note_type": "task_type", "note": str(plan.get("task_type") or "")},
        {"note_type": "method", "note": "Rows are normalized from retrieved variables only; no new data was ingested and no values were invented."},
        {"note_type": "method", "note": "Numeric values are extracted from existing evidence quotes when present; arithmetic is performed only after comparability validation."},
        {"note_type": "comparability", "note": str(comparability.get("status") or "insufficient_metadata")},
    ]
    for issue in comparability.get("issues") or []:
        notes.append({"note_type": "comparability_issue", "note": issue})
    return notes


def data_gaps(evidence_packet: dict[str, Any], comparability: dict[str, Any]) -> list[dict[str, str]]:
    gaps = []
    if not evidence_packet.get("variables"):
        gaps.append({"gap_type": "variables", "description": "No structured variables were retrieved."})
    for field_name, label in [("geography", "geography"), ("time_period", "time coverage"), ("unit", "unit"), ("value", "numeric value")]:
        missing = [item.get("metric_name") or "Unnamed metric" for item in evidence_packet.get("variables") or [] if item.get(field_name) in (None, "")]
        if missing:
            gaps.append({"gap_type": field_name, "description": f"Missing {label} for {len(missing)} variable(s): {', '.join(missing[:5])}."})
    for issue in comparability.get("issues") or []:
        gaps.append({"gap_type": "comparability", "description": issue})
    return gaps or [{"gap_type": "none", "description": "No obvious metadata gaps detected in retrieved rows."}]
