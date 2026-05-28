from __future__ import annotations

from openpyxl import load_workbook

from app.services.research_task import (
    AnswerSynthesizer,
    ComparabilityValidator,
    EvidencePacketBuilder,
    ResearchTaskPlanner,
    TableExcelExportService,
    execute_research_task,
    extract_numeric_value,
)


def sample_results() -> dict:
    return {
        "closest_variables": [
            {
                "object_id": "var-1",
                "title": "Singapore startup funding by stage",
                "score": 0.84,
                "definition": "Funding amount split by investment stage.",
                "unit": "USD million",
                "availability": "obtainable",
                "temporal_coverage": "2020-2023",
                "geographic_coverage": "Singapore",
                "source_url": "https://example.org/report",
                "evidence_quote": "Seed funding reached 42.5 USD million in 2023.",
                "data_source": "Singapore Venture Report",
            },
            {
                "object_id": "var-2",
                "title": "Singapore startup deal count by stage",
                "score": 0.67,
                "definition": "Number of deals split by investment stage.",
                "unit": "deals",
                "availability": "private",
                "temporal_coverage": "2020-2023",
                "geographic_coverage": "Singapore",
                "source_url": "https://example.org/report",
                "evidence_quote": "Seed stage had 120 deals in 2023.",
                "data_source": "Singapore Venture Report",
            },
        ],
        "relevant_reports": [
            {
                "object_id": "report-1",
                "title": "Singapore Venture Report",
                "publisher": "Example Publisher",
                "geography": "Singapore",
                "report_year": 2024,
                "source_url": "https://example.org/report",
                "score": 0.9,
            }
        ],
        "source_links": [{"title": "Report URL", "source_url": "https://example.org/report", "availability": "obtainable"}],
        "relevant_organizations": [{"object_id": "org-1", "name": "Startup SG", "geography": "Singapore", "score": 0.5}],
        "limitations": [],
    }


def test_research_task_planner_classifies_create_excel() -> None:
    plan = ResearchTaskPlanner().plan("Create an Excel of Singapore startup funding by stage")

    assert plan.task_type == "create_excel"
    assert plan.geography == "Singapore"
    assert plan.domain == "startup funding"
    assert plan.dimension == "stage"
    assert plan.output_format == "xlsx"


def test_evidence_packet_creation() -> None:
    plan = ResearchTaskPlanner().plan("Create an Excel of Singapore startup funding by stage")
    packet = EvidencePacketBuilder().build("Create an Excel of Singapore startup funding by stage", plan, sample_results())

    assert packet["query"]
    assert packet["variables"][0]["metric_name"] == "Singapore startup funding by stage"
    assert packet["variables"][0]["value"] == 42.5
    assert packet["source_urls"] == ["https://example.org/report"]
    assert "private" in packet["availability_labels"]
    assert packet["geography_coverage"] == ["Singapore"]


def test_answer_synthesis_prompt_formatting() -> None:
    plan = ResearchTaskPlanner().plan("Singapore startup funding by stage")
    packet = EvidencePacketBuilder().build("Singapore startup funding by stage", plan, sample_results())
    synthesizer = AnswerSynthesizer()
    prompt = synthesizer.prompt_payload(packet)
    answer = synthesizer.synthesize(packet)

    assert "Lead with a direct answer" in prompt["instruction"]
    assert "Direct answer:" in answer
    assert "Direct matches:" in answer
    assert "Availability:" in answer
    assert "Singapore Venture Report" in answer


def test_numeric_value_extraction_from_evidence() -> None:
    value, unit = extract_numeric_value("Funding was $1,234.5 million in 2023.")

    assert value == 1234.5
    assert unit == "$"


def test_excel_export_sheet_structure(tmp_path) -> None:
    plan = ResearchTaskPlanner().plan("Create an Excel of Singapore startup funding by stage")
    packet = EvidencePacketBuilder().build("Create an Excel of Singapore startup funding by stage", plan, sample_results())
    rows = TableExcelExportService().build_rows(packet)
    comparability = ComparabilityValidator().validate(rows)
    export = TableExcelExportService().export(packet, output_dir=tmp_path, output_format="xlsx", comparability=comparability)

    wb = load_workbook(export["path"])
    assert set(wb.sheetnames) == {
        "normalized_data",
        "source_variables",
        "source_reports",
        "methodology_notes",
        "data_gaps",
    }
    header = [cell.value for cell in wb["normalized_data"][1]]
    assert "metric_name" in header
    assert "comparability_status" in header


def test_comparability_blocks_unsafe_aggregation() -> None:
    rows = [
        {"metric_name": "Funding amount", "geography": "Singapore", "time_period": "2023", "unit": "USD", "dimension": "stage", "source_url": "a", "availability": "obtainable"},
        {"metric_name": "Funding amount", "geography": "Singapore", "time_period": "2022", "unit": "SGD", "dimension": "stage", "source_url": "b", "availability": "private"},
    ]

    result = ComparabilityValidator().validate(rows, aggregation_requested=True)

    assert result["status"] == "not_comparable"
    assert result["can_aggregate"] is False
    assert any("Mixed unit" in issue for issue in result["issues"])


def test_execute_research_task_uses_tool_caller_without_llm(tmp_path) -> None:
    calls = []

    def fake_tool(name, args):
        calls.append((name, args))
        return {"ok": True, "data": sample_results()}

    result = execute_research_task(
        "Create an Excel of Singapore startup funding by stage",
        tool_caller=fake_tool,
        output_dir=tmp_path,
        output_format="xlsx",
    )

    assert result["ok"] is True
    assert calls[0][0] == "find_data"
    assert result["export"]["format"] == "xlsx"
    assert result["answer"].startswith("Direct answer:")
